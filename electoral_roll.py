# """
# Electoral Roll OCR Extractor — Tamil Nadu (Image-based PDFs)
# ============================================================
# Strategy: The PDF pages are scanned images. Each page has a 3-column layout
# where each card group (row of 3 cards) produces these OCR lines:

#   Line A: "[serial] voter_id  [serial] voter_id  [serial] voter_id"
#   Line B: "Name : PERSON1  Name : PERSON2  Name : PERSON3"
#   Line C: "Father/Husband Name: X  Father/Husband Name: Y  Father/Husband Name: Z"
#   Line D: "House Number: h1 Photo House Number: h2 Photo House Number: h3 Photo"
#   Line E: "Age: n1 Gender: G1  Age: n2 Gender: G2  Age: n3 Gender: G3"

# We parse row-by-row, splitting each field line into 3 columns.

# NEW in this version:
#   1. Reads summary page first to get expected voter count.
#   2. Compares extracted count vs expected; cascading reruns if >2% off.
#      (2% threshold: 20 missed voters per 1000 is already unacceptable.)
#   3. Guaranteed age & gender extraction with multiple fallback patterns.
#   4. Saves one CSV per sheet (All Records + per-section) inside a single ZIP,
#      AND also writes an .xlsx with named sheets for convenience.

# Usage:
#     python electoral_roll.py <pdf_path> [--start-page N] [--end-page N] [--output out.xlsx]
# """

# import re
# import sys
# import csv
# import zipfile
# import argparse
# import logging
# from io import StringIO
# from pathlib import Path
# from dataclasses import dataclass, asdict, fields
# from typing import Optional

# import pytesseract
# from PIL import Image
# from pdf2image import convert_from_path
# import openpyxl
# from openpyxl.styles import Font, PatternFill, Alignment

# logging.basicConfig(
#     level=logging.INFO,
#     format="%(asctime)s  %(levelname)-8s  %(message)s",
#     datefmt="%H:%M:%S",
# )
# log = logging.getLogger(__name__)

# TESSERACT_CONFIG     = "--psm 6 --oem 3"
# TESSERACT_CONFIG_ALT = "--psm 4 --oem 3"   # alternative for retry

# DPI_DEFAULT = 200
# # Cascade retry schedule: each tuple is (dpi, tesseract_config).
# # Tried in order until count is within threshold or list is exhausted.
# RETRY_SCHEDULE = [
#     (250, "--psm 6 --oem 3"),   # pass 2: slightly higher DPI, same layout
#     (300, "--psm 4 --oem 3"),   # pass 3: high DPI, single-column mode
#     (350, "--psm 6 --oem 1"),   # pass 4: max DPI, LSTM-only engine
# ]
# MISMATCH_THRESHOLD = 0.02       # 2% — 20 missed voters per 1000 is already significant

# # Windows: set Tesseract path if not in system PATH
# pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"


# @dataclass
# class VoterRecord:
#     sno: str = ""
#     voter_id: str = ""
#     name: str = ""
#     father_husband_name: str = ""
#     house_number: str = ""
#     age: str = ""
#     gender: str = ""
#     part_no: str = ""
#     ward_no: str = ""
#     street_area_name: str = ""
#     constituency_no: str = ""
#     constituency_name: str = ""

#     def is_valid(self) -> bool:
#         return bool(self.name.strip())


# # ── Summary page parser ────────────────────────────────────────────────────────

# def parse_summary_count(pdf_path: str, total_pages: int, poppler_path) -> Optional[int]:
#     """
#     Render only the last page (summary/statistics page) and extract the total
#     voter count.  Tamil Nadu roll summaries typically contain a line like:
#       "Total Electors : 1234"  or  "Total Voters : 1234"  or
#       "குடிமக்கள் மொத்தம் : 1234"
#     Returns None if nothing is found.
#     """
#     log.info("Reading summary page to get expected voter count …")
#     try:
#         images = convert_from_path(
#             pdf_path, dpi=DPI_DEFAULT,
#             first_page=total_pages, last_page=total_pages,
#             fmt="jpeg", poppler_path=poppler_path,
#         )
#         if not images:
#             return None
#         text = pytesseract.image_to_string(images[0], config=TESSERACT_CONFIG)
#         # Try multiple patterns used in Tamil Nadu rolls
#         patterns = [
#             r"Total\s+Electors?\s*[:\-=]\s*([\d,]+)",
#             r"Total\s+Voters?\s*[:\-=]\s*([\d,]+)",
#             r"(?:Grand\s+)?Total\s*[:\-=]\s*([\d,]+)",
#             r"மொத்தம்\s*[:\-=]\s*([\d,]+)",       # Tamil script
#             r"Total\s+\d+\s+\d+\s+\d+\s+([\d,]+)", # tabular summary row
#         ]
#         for pat in patterns:
#             m = re.search(pat, text, re.IGNORECASE)
#             if m:
#                 count = int(m.group(1).replace(",", ""))
#                 log.info(f"Summary page → expected voter count: {count}")
#                 return count
#         log.warning("Could not find voter count on summary page; validation skipped.")
#         log.debug(f"Summary OCR text:\n{text[:800]}")
#         return None
#     except Exception as e:
#         log.warning(f"Summary page read failed: {e}")
#         return None


# # ── Header parser ──────────────────────────────────────────────────────────────

# def parse_page_header(lines: list) -> dict:
#     ctx = {
#         "constituency_no": "", "constituency_name": "",
#         "part_no": "", "section_no": "",
#         "ward_no": "", "street_area_name": "",
#     }
#     for line in lines[:8]:
#         m = re.search(
#             r"Assembly\s+Constituency\s+No\s+and\s+Name\s*[:=]\s*"
#             r"(\d+)[^\w]+([\w\s]+?)\s+Part\s+No\.?\s*[:\s]*(\d+)",
#             line, re.IGNORECASE
#         )
#         if m:
#             ctx["constituency_no"]   = m.group(1).strip()
#             ctx["constituency_name"] = m.group(2).strip()
#             ctx["part_no"]           = m.group(3).strip()

#         m2 = re.search(r"Part\s+No\.?\s*[:\s]*(\d+)", line, re.IGNORECASE)
#         if m2 and not ctx["part_no"]:
#             ctx["part_no"] = m2.group(1).strip()

#         m3 = re.search(r"Section\s+No\s+and\s+Name\s+(\d+)[^\w]+(.*)", line, re.IGNORECASE)
#         if m3:
#             ctx["section_no"]      = m3.group(1).strip()
#             raw                    = m3.group(2).strip()
#             ctx["street_area_name"] = raw
#             ward_m = re.search(r"Ward\s+([\w\s]+?)(?:,|$)", raw, re.IGNORECASE)
#             if ward_m:
#                 ctx["ward_no"] = ward_m.group(1).strip()

#     return ctx


# # ── Field line splitters ───────────────────────────────────────────────────────

# def split_name_line(line: str) -> list:
#     all_matches = list(re.finditer(
#         r"\bName\s*[:\-!=+}\|;*?#\ufffd]|\bName\s+[a-z](?=\s)",
#         line, re.IGNORECASE
#     ))
#     if not all_matches:
#         return []
#     filtered = []
#     for m in all_matches:
#         before = line[:m.start()].rstrip()
#         if re.search(r"(?:Father|Husband|Mother)\s*$", before, re.IGNORECASE):
#             continue
#         filtered.append(m)
#     if not filtered:
#         return []
#     results = []
#     for i, m in enumerate(filtered):
#         start = m.end()
#         end   = filtered[i + 1].start() if i + 1 < len(filtered) else len(line)
#         val   = line[start:end].strip().rstrip("-~").strip()
#         val   = re.sub(r"\s+Photo\s*$", "", val, flags=re.IGNORECASE).strip()
#         val   = re.sub(r"\s+(?:Father|Husband|Mother)$", "", val, flags=re.IGNORECASE).strip()
#         if val:
#             results.append(val)
#     return results


# def split_relation_line(line: str) -> list:
#     matches = list(re.finditer(
#         r"(?:Father|Husband|Mother)\s+(?:Name|Narne)\s*[:\-!=?+]\s*",
#         line, re.IGNORECASE
#     ))
#     if not matches:
#         return []
#     results = []
#     for i, m in enumerate(matches):
#         start = m.end()
#         end   = matches[i + 1].start() if i + 1 < len(matches) else len(line)
#         val   = line[start:end].strip().rstrip("-~").strip()
#         val   = re.sub(r"\s+Photo\s*$", "", val, flags=re.IGNORECASE).strip()
#         if val:
#             results.append(val)
#     return results


# def split_house_line(line: str) -> list:
#     matches = list(re.finditer(r"House\s+Number\s*[:\-!=+?]\s*", line, re.IGNORECASE))
#     if not matches:
#         return []
#     results = []
#     for i, m in enumerate(matches):
#         start = m.end()
#         end   = matches[i + 1].start() if i + 1 < len(matches) else len(line)
#         val   = line[start:end].strip()
#         val   = re.split(r"\s+Photo\b|\s{3,}", val)[0].strip()
#         if val:
#             results.append(val)
#     return results


# def split_age_gender_line(line: str) -> list:
#     """
#     Split: "Age : 83 Gender : Male  Age + 38 Gender : Male  Age : 36 Gender : Male"
#     Returns list of (age, gender) tuples.

#     Multiple fallback patterns handle OCR garbling such as:
#       - "Age: 83 Sex: Male"
#       - "Age 83 Gender Male"
#       - "A9e : 25 Gender : Female"   (OCR '9' for 'g')
#     """
#     def _norm_gender(g: str) -> str:
#         g = g.strip().title()
#         lg = g.lower()
#         if lg in ("femaie", "femail", "fernale"):
#             return "Female"
#         if lg in ("mate", "mal", "mae"):
#             return "Male"
#         if "third" in lg:
#             return "Third Gender"
#         return g

#     results = []

#     # Pattern 1: canonical form — Age : N Gender : G
#     p1 = re.compile(
#         r"(?:Age|A9e|Ag[e3])\s*[:\-!+?=]?\s*(\d{1,3}\.?)\s+"
#         r"(?:Gender|Sex|Gen[d]er)\s*[:\-!+?=]\s*"
#         r"(Male|Mate|Mae|Mal|Female|Femaie|Femail|Fernale|Third\s*Gender)",
#         re.IGNORECASE
#     )
#     for m in p1.finditer(line):
#         results.append((m.group(1).rstrip("."), _norm_gender(m.group(2))))

#     if results:
#         return results

#     # Pattern 2: Age and Gender on separate sub-tokens (OCR splits lines)
#     # e.g. "Age : 34" on one captured group, "Gender : Male" somewhere later
#     ages    = re.findall(r"(?:Age|A9e)\s*[:\-!+?=]?\s*(\d{1,3})", line, re.IGNORECASE)
#     genders = re.findall(
#         r"(?:Gender|Sex)\s*[:\-!+?=]\s*(Male|Mate|Mae|Mal|Female|Femaie|Femail|Third\s*Gender)",
#         line, re.IGNORECASE
#     )
#     if ages and genders:
#         for i in range(min(len(ages), len(genders))):
#             results.append((ages[i], _norm_gender(genders[i])))
#         return results

#     # Pattern 3: bare numbers after Age keyword (last resort)
#     age_only = re.findall(r"(?:Age|A9e)\s*[:\-!+?=]?\s*(\d{1,3})", line, re.IGNORECASE)
#     if age_only:
#         # Gender can't be determined — return with empty string so age is preserved
#         return [(a, "") for a in age_only]

#     return []


# def extract_voter_ids_from_serial_line(line: str) -> list:
#     pattern = re.compile(r"(?i)([A-Za-z]{2,4})([0-9]{5,10})")
#     return [(m.group(1) + m.group(2)).upper() for m in pattern.finditer(line)]


# def extract_serials_from_line(line: str) -> list:
#     return re.findall(r"(?:^|[|\[\(#\s])(\d{1,4})(?:[\]\)]|\s|$)", line)


# def is_serial_line(line: str) -> bool:
#     blocked = ["name", "father", "husband", "mother", "age", "gender",
#                "house", "available", "photo", "assembly", "section",
#                "constituency", "part no", "ward", "age as on"]
#     ll = line.lower()
#     if any(b in ll for b in blocked):
#         return False
#     if not re.search(r"\d", line):
#         return False
#     return bool(
#         re.search(r"[A-Za-z]{2,}\d{3,}", line) or
#         re.search(r"\[\d+\]", line) or
#         re.search(r"^\s*\d{1,3}\s*$", line)
#     )


# # ── Group parser ───────────────────────────────────────────────────────────────

# def _parse_group(grp_lines: list, ctx: dict) -> list:
#     """Parse a single voter-card group (1–3 cards) and return VoterRecords."""
#     voter_ids, serials          = [], []
#     names, fathers, houses      = [], [], []
#     age_genders: list           = []

#     for gl in grp_lines:
#         gll = gl.lower()
#         if is_serial_line(gl):
#             voter_ids = extract_voter_ids_from_serial_line(gl)
#             serials   = extract_serials_from_line(gl)
#             continue

#         if "father name" in gll or "husband name" in gll or "mother name" in gll:
#             f = split_relation_line(gl)
#             if f:
#                 fathers = f

#         if re.search(r"\bname\s*[:\-!=+}\|;*?#\ufffd]|\bname\s+[a-z]\s", gll):
#             n = split_name_line(gl)
#             if n:
#                 names = n

#         if "house number" in gll:
#             h = split_house_line(gl)
#             if h:
#                 houses = h

#         if re.search(r"\bage\b|\ba9e\b", gll):
#             ag = split_age_gender_line(gl)
#             if ag:
#                 age_genders.extend(ag)

#     # De-duplicate age_genders (same entry can appear twice if line repeated)
#     seen = set()
#     deduped_ag = []
#     for entry in age_genders:
#         if entry not in seen:
#             seen.add(entry)
#             deduped_ag.append(entry)
#     age_genders = deduped_ag

#     records   = []
#     n_cards   = max(len(names), len(voter_ids), 1)
#     for k in range(min(n_cards, 3)):
#         rec = VoterRecord(
#             part_no=ctx.get("part_no", ""),
#             ward_no=ctx.get("ward_no", ""),
#             street_area_name=ctx.get("street_area_name", ""),
#             constituency_no=ctx.get("constituency_no", ""),
#             constituency_name=ctx.get("constituency_name", ""),
#         )
#         rec.sno                 = serials[k]    if k < len(serials)     else ""
#         rec.voter_id            = voter_ids[k]  if k < len(voter_ids)   else ""
#         rec.name                = names[k]      if k < len(names)       else ""
#         rec.father_husband_name = fathers[k]    if k < len(fathers)     else ""
#         rec.house_number        = houses[k]     if k < len(houses)      else ""
#         if k < len(age_genders):
#             rec.age, rec.gender = age_genders[k]
#         if rec.is_valid():
#             records.append(rec)
#     return records


# # ── Page processor ─────────────────────────────────────────────────────────────

# def process_page(image, page_num: int, tess_config: str = TESSERACT_CONFIG) -> tuple:
#     """OCR a page and extract VoterRecords."""
#     log.info(f"  OCR page {page_num} …")
#     text  = pytesseract.image_to_string(image, config=tess_config)
#     lines = [l.strip() for l in text.splitlines()]

#     non_empty   = [l for l in lines if l]
#     ctx         = parse_page_header(non_empty[:10])
#     log.info(
#         f"    → {ctx['constituency_no']}-{ctx['constituency_name']} "
#         f"| Part {ctx['part_no']} | Ward: {ctx['ward_no']}"
#     )

#     if not any(kw in text for kw in ["Name :", "Name:", "Age :", "Age:", "Gender :"]):
#         log.info("    → Skip (no voter fields detected)")
#         return ctx, []

#     header_keywords = ["assembly constituency", "section no and name", "age as on"]
#     current_ctx     = dict(ctx)
#     records         = []
#     current_group   = []

#     for line in non_empty:
#         ll = line.lower()

#         if "section no and name" in ll:
#             new_ctx = parse_page_header([line])
#             for k, v in new_ctx.items():
#                 if v:
#                     current_ctx[k] = v
#             continue

#         if any(kw in ll for kw in header_keywords):
#             continue

#         if re.search(r'\bAvailable\b', line, re.IGNORECASE):
#             if re.search(r'\bAge\b.*\bGender\b', line, re.IGNORECASE):
#                 current_group.append(line)
#             if current_group:
#                 records.extend(_parse_group(current_group, dict(current_ctx)))
#                 current_group = []
#             continue

#         current_group.append(line)

#     if current_group:
#         records.extend(_parse_group(current_group, dict(current_ctx)))

#     # ── Post-pass: recover missing age/gender via a second targeted OCR ────────
#     missing = [r for r in records if not r.age or not r.gender]
#     if missing:
#         log.info(f"    → {len(missing)} records missing age/gender; running targeted re-OCR …")
#         records = _recover_age_gender(image, records, tess_config)

#     log.info(f"    → {len(records)} records extracted")
#     return ctx, records


# def _recover_age_gender(image, records: list, tess_config: str) -> list:
#     """
#     Re-OCR the image with --psm 4 (single-column assumption) to pick up
#     Age/Gender lines that the main pass missed, and patch records in-place.
#     Works by scanning for age-gender patterns in the alternative OCR output
#     and matching them positionally to records without age/gender.
#     """
#     alt_text = pytesseract.image_to_string(image, config=TESSERACT_CONFIG_ALT)
#     all_ag   = []
#     for line in alt_text.splitlines():
#         ag = split_age_gender_line(line.strip())
#         all_ag.extend(ag)

#     missing_idx = [i for i, r in enumerate(records) if not r.age or not r.gender]
#     fill_ag     = [(a, g) for a, g in all_ag if a]  # only entries with real age

#     if not fill_ag:
#         return records

#     # Pair sequentially — OCR order matches card order on page
#     for j, idx in enumerate(missing_idx):
#         if j < len(fill_ag):
#             age, gender = fill_ag[j]
#             if not records[idx].age:
#                 records[idx].age = age
#             if not records[idx].gender:
#                 records[idx].gender = gender

#     return records


# # ── Validation + retry ─────────────────────────────────────────────────────────

# def run_extraction(pdf_path: str, start_page: int, end_page: int, dpi: int,
#                    poppler_path, tess_config: str = TESSERACT_CONFIG) -> tuple:
#     """
#     Render pages and extract records.  Returns (all_records, section_map, current_ctx).
#     """
#     log.info(f"Rendering pages {start_page}–{end_page} at DPI={dpi} …")
#     images = convert_from_path(
#         str(pdf_path), dpi=dpi,
#         first_page=start_page, last_page=end_page,
#         fmt="jpeg", thread_count=2,
#         poppler_path=poppler_path,
#     )
#     log.info(f"Rendered {len(images)} page(s)")

#     all_records = []
#     section_map: dict = {}
#     current_ctx: dict = {}

#     for idx, image in enumerate(images):
#         ctx, records = process_page(image, start_page + idx, tess_config)
#         for k, v in ctx.items():
#             if v:
#                 current_ctx[k] = v

#         for rec in records:
#             if not rec.constituency_no:   rec.constituency_no   = current_ctx.get("constituency_no", "")
#             if not rec.constituency_name: rec.constituency_name = current_ctx.get("constituency_name", "")
#             if not rec.part_no:           rec.part_no           = current_ctx.get("part_no", "")
#             if not rec.ward_no:           rec.ward_no           = current_ctx.get("ward_no", "")
#             if not rec.street_area_name:  rec.street_area_name  = current_ctx.get("street_area_name", "")

#         all_records.extend(records)
#         for rec in records:
#             key       = rec.street_area_name or f"Section_{current_ctx.get('section_no', '?')}"
#             sheet_key = (key[:28] + "..") if len(key) > 30 else key
#             section_map.setdefault(sheet_key, []).append(rec)

#     return all_records, section_map, current_ctx


# # ── Output writers ─────────────────────────────────────────────────────────────

# HEADER_COLS = [
#     "S.NO", "VOTER_ID", "NAME", "FATHER_HUSBAND_NAME",
#     "HOUSE_NUMBER", "AGE", "GENDER",
#     "PART_NO", "WARD_NO", "STREET_AREA_NAME",
#     "CONSTITUENCY_NO", "CONSTITUENCY_NAME",
# ]
# HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
# HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
# ALT_FILL    = PatternFill("solid", fgColor="D6E4F0")


# def _record_row(rec: VoterRecord) -> list:
#     d = asdict(rec)
#     return [
#         d["sno"], d["voter_id"], d["name"], d["father_husband_name"],
#         d["house_number"], d["age"], d["gender"],
#         d["part_no"], d["ward_no"], d["street_area_name"],
#         d["constituency_no"], d["constituency_name"],
#     ]


# def _write_xlsx_sheet(ws, records: list):
#     ws.append(HEADER_COLS)
#     for cell in ws[1]:
#         cell.fill      = HEADER_FILL
#         cell.font      = HEADER_FONT
#         cell.alignment = Alignment(horizontal="center")

#     for i, rec in enumerate(records, start=2):
#         ws.append(_record_row(rec))
#         if i % 2 == 0:
#             for cell in ws[i]:
#                 cell.fill = ALT_FILL

#     widths = {"A":6,"B":14,"C":28,"D":28,"E":14,"F":6,"G":10,"H":8,"I":18,"J":50,"K":14,"L":22}
#     for col, w in widths.items():
#         ws.column_dimensions[col].width = w
#     ws.freeze_panes = "A2"


# def save_to_xlsx(all_records, section_map, output_path: str):
#     wb      = openpyxl.Workbook()
#     ws_all  = wb.active
#     ws_all.title = "All Records"
#     _write_xlsx_sheet(ws_all, all_records)
#     log.info(f"  Sheet 'All Records' → {len(all_records)} rows")

#     for section_key, recs in section_map.items():
#         safe = re.sub(r"[\\/*?:\[\]]", "", section_key)[:31]
#         ws   = wb.create_sheet(title=safe)
#         _write_xlsx_sheet(ws, recs)
#         log.info(f"  Sheet '{safe}' → {len(recs)} rows")

#     wb.save(output_path)
#     log.info(f"Excel saved → {output_path}")


# def save_to_csv_zip(all_records, section_map, zip_path: str):
#     """
#     Save one CSV per sheet inside a ZIP file.
#     Sheets:
#       - All_Records.csv
#       - <section_name>.csv  for each section
#     """
#     def _csv_bytes(records: list) -> bytes:
#         buf = StringIO()
#         w   = csv.writer(buf)
#         w.writerow(HEADER_COLS)
#         for rec in records:
#             w.writerow(_record_row(rec))
#         return buf.getvalue().encode("utf-8-sig")   # utf-8-sig = Excel-friendly BOM

#     with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
#         zf.writestr("All_Records.csv", _csv_bytes(all_records))
#         log.info(f"  CSV 'All_Records.csv' → {len(all_records)} rows")

#         for section_key, recs in section_map.items():
#             safe_name = re.sub(r"[\\/*?:\[\]<>|\"']", "_", section_key)[:60]
#             fname     = f"{safe_name}.csv"
#             zf.writestr(fname, _csv_bytes(recs))
#             log.info(f"  CSV '{fname}' → {len(recs)} rows")

#     log.info(f"CSV ZIP saved → {zip_path}")


# # ── CLI ────────────────────────────────────────────────────────────────────────

# def main():
#     parser = argparse.ArgumentParser(
#         description="Extract voter data from Tamil Nadu Electoral Roll PDFs"
#     )
#     parser.add_argument("pdf_path")
#     parser.add_argument("--start-page", type=int, default=3,
#         help="First page (1-based). Default=3 skips cover + map.")
#     parser.add_argument("--end-page", type=int, default=None,
#         help="Last page (1-based). Default=second-to-last (skips summary page).")
#     parser.add_argument("--dpi", type=int, default=DPI_DEFAULT,
#         help=f"Rendering DPI (default={DPI_DEFAULT}). Higher = better but slower.")
#     parser.add_argument("--output", default=None,
#         help="Output .xlsx path (a matching .zip of CSVs is also written)")
#     parser.add_argument("--no-retry", action="store_true",
#         help="Disable automatic cascading retry when count mismatches > 2%%")
#     args = parser.parse_args()

#     pdf_path = Path(args.pdf_path)
#     if not pdf_path.exists():
#         log.error(f"Not found: {pdf_path}")
#         sys.exit(1)

#     from pypdf import PdfReader
#     total_pages = len(PdfReader(str(pdf_path)).pages)
#     start_page  = args.start_page
#     end_page    = args.end_page or (total_pages - 1)
#     xlsx_path   = args.output or str(pdf_path.with_suffix(".xlsx"))
#     csv_zip     = str(Path(xlsx_path).with_suffix(".csv.zip"))

#     log.info(f"PDF: {pdf_path.name}  ({total_pages} pages)")
#     log.info(f"Processing pages {start_page}–{end_page}")

#     # Poppler path for Windows (winget install location)
#     import os
#     _poppler = r"C:\Users\sarat\AppData\Local\Microsoft\WinGet\Packages\oschwartz10612.Poppler_Microsoft.Winget.Source_8wekyb3d8bbwe\poppler-25.07.0\Library\bin"
#     poppler_path = _poppler if os.path.isdir(_poppler) else None

#     # ── Step 1: read summary to get expected count ──────────────────────────
#     expected_count = parse_summary_count(str(pdf_path), total_pages, poppler_path)

#     # ── Step 2: first extraction pass ──────────────────────────────────────
#     all_records, section_map, current_ctx = run_extraction(
#         str(pdf_path), start_page, end_page, args.dpi, poppler_path
#     )
#     log.info(f"\nFirst pass → {len(all_records)} records extracted")

#     # ── Step 3: validate count; cascade retries until within 2% ────────────────
#     if expected_count and not args.no_retry:
#         best_records    = all_records
#         best_section    = section_map
#         best_gap        = abs(len(all_records) - expected_count)

#         for attempt, (retry_dpi, retry_cfg) in enumerate(RETRY_SCHEDULE, start=2):
#             extracted   = len(best_records)
#             diff_ratio  = abs(extracted - expected_count) / max(expected_count, 1)
#             missed      = expected_count - extracted

#             if diff_ratio <= MISMATCH_THRESHOLD:
#                 log.info(
#                     f"Count OK after pass {attempt - 1}: "
#                     f"{extracted}/{expected_count} ({diff_ratio:.2%} — within {MISMATCH_THRESHOLD:.0%} threshold)."
#                 )
#                 break

#             log.warning(
#                 f"Pass {attempt - 1}: extracted {extracted} / expected {expected_count} "
#                 f"({diff_ratio:.2%} off = ~{missed} voters missing). "
#                 f"Starting pass {attempt} at DPI={retry_dpi}, config='{retry_cfg}' …"
#             )

#             retry_records, retry_section_map, _ = run_extraction(
#                 str(pdf_path), start_page, end_page,
#                 retry_dpi, poppler_path, tess_config=retry_cfg
#             )
#             retry_count = len(retry_records)
#             log.info(f"Pass {attempt} → {retry_count} records")

#             # ── Merge strategy ────────────────────────────────────────────────
#             # Don't just pick one pass — merge both, deduplicating by voter_id.
#             # Records with no voter_id are kept from whichever pass had more.
#             # This way a voter captured in pass 1 but missed in pass 2 survives.
#             merged_by_id: dict = {}
#             no_id_records: list = []

#             def _absorb(records):
#                 for r in records:
#                     if r.voter_id:
#                         existing = merged_by_id.get(r.voter_id)
#                         if existing is None:
#                             merged_by_id[r.voter_id] = r
#                         else:
#                             # Fill in any missing fields from the other pass
#                             if not existing.age    and r.age:    existing.age    = r.age
#                             if not existing.gender and r.gender: existing.gender = r.gender
#                             if not existing.name   and r.name:   existing.name   = r.name
#                             if not existing.father_husband_name and r.father_husband_name:
#                                 existing.father_husband_name = r.father_husband_name
#                     else:
#                         no_id_records.append(r)

#             _absorb(best_records)
#             _absorb(retry_records)

#             # For no-id records: take the larger set (better-quality pass)
#             if len(retry_records) > len(best_records):
#                 no_id_from = [r for r in retry_records if not r.voter_id]
#             else:
#                 no_id_from = [r for r in best_records  if not r.voter_id]

#             merged = list(merged_by_id.values()) + no_id_from
#             merged_gap = abs(len(merged) - expected_count)

#             log.info(
#                 f"After merge: {len(merged)} records "
#                 f"(gap {merged_gap} vs expected {expected_count})"
#             )

#             # Always take the merged result if it's at least as good as best
#             if merged_gap <= best_gap:
#                 best_gap        = merged_gap
#                 best_records    = merged
#                 # Rebuild section_map from merged records
#                 best_section = {}
#                 for rec in merged:
#                     key       = rec.street_area_name or f"Section_{current_ctx.get('section_no', '?')}"
#                     sheet_key = (key[:28] + "..") if len(key) > 30 else key
#                     best_section.setdefault(sheet_key, []).append(rec)
#             else:
#                 log.info("Merged result is worse than best so far; keeping previous best.")

#         else:
#             # Loop exhausted without hitting threshold
#             final_gap   = abs(len(best_records) - expected_count)
#             final_ratio = final_gap / max(expected_count, 1)
#             log.warning(
#                 f"All {len(RETRY_SCHEDULE) + 1} passes done. "
#                 f"Best result: {len(best_records)}/{expected_count} "
#                 f"({final_ratio:.2%} gap = ~{final_gap} voters still missing). "
#                 f"Saving best available data."
#             )

#         all_records = best_records
#         section_map = best_section

#     elif expected_count and args.no_retry:
#         diff_ratio = abs(len(all_records) - expected_count) / max(expected_count, 1)
#         log.info(
#             f"Retry disabled. Extracted {len(all_records)}/{expected_count} "
#             f"({diff_ratio:.2%} gap)."
#         )
#     else:
#         log.info("No expected count found on summary page; skipping validation.")

#     final_count = len(all_records)
#     age_missing    = sum(1 for r in all_records if not r.age)
#     gender_missing = sum(1 for r in all_records if not r.gender)
#     log.info(
#         f"\nFinal totals: {final_count} records | "
#         f"{len(section_map)} sections | "
#         f"missing age: {age_missing} | missing gender: {gender_missing}"
#     )

#     if not all_records:
#         log.warning("No records extracted. Check --start-page / --end-page.")
#         sys.exit(1)

#     # ── Step 4: save outputs ────────────────────────────────────────────────
#     save_to_xlsx(all_records, section_map, xlsx_path)
#     save_to_csv_zip(all_records, section_map, csv_zip)

#     log.info(f"\n✅  Done")
#     log.info(f"   Excel  → {xlsx_path}")
#     log.info(f"   CSVs   → {csv_zip}  (one CSV per sheet inside ZIP)")


# if __name__ == "__main__":
#     main()







"""
Electoral Roll OCR Extractor — Tamil Nadu (Image-based PDFs)
============================================================
Strategy: The PDF pages are scanned images. Each page has a 3-column layout
where each card group (row of 3 cards) produces these OCR lines:

  Line A: "[serial] voter_id  [serial] voter_id  [serial] voter_id"
  Line B: "Name : PERSON1  Name : PERSON2  Name : PERSON3"
  Line C: "Father/Husband Name: X  Father/Husband Name: Y  Father/Husband Name: Z"
  Line D: "House Number: h1 Photo House Number: h2 Photo House Number: h3 Photo"
  Line E: "Age: n1 Gender: G1  Age: n2 Gender: G2  Age: n3 Gender: G3"

We parse row-by-row, splitting each field line into 3 columns.

FIXES in this version:
  1. Ward No extraction — now matches "Ward No. : 5", "Ward : 5", "Ward No 5"
     AND the embedded "Ward XYZ" inside a section name string.
  2. ZIP file is no longer created (xlsx only).
  3. --batch mode: output xlsx sheet is named after the PDF filename stem.

Usage:
    python electoral_roll.py <pdf_path> [--start-page N] [--end-page N] [--output out.xlsx]
    python electoral_roll.py <pdf_path> --batch          # sheet named after PDF filename
"""

import re
import sys
import csv
import argparse
import logging
from io import StringIO
from pathlib import Path
from dataclasses import dataclass, asdict, fields
from typing import Optional

import pytesseract
from PIL import Image
from pdf2image import convert_from_path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

TESSERACT_CONFIG     = "--psm 6 --oem 3"
TESSERACT_CONFIG_ALT = "--psm 4 --oem 3"   # alternative for retry

DPI_DEFAULT = 200
# Cascade retry schedule: each tuple is (dpi, tesseract_config).
RETRY_SCHEDULE = [
    (250, "--psm 6 --oem 3"),
    (300, "--psm 4 --oem 3"),
    (350, "--psm 6 --oem 1"),
]
MISMATCH_THRESHOLD = 0.02

# Windows: set Tesseract path if not in system PATH
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"


@dataclass
class VoterRecord:
    sno: str = ""
    voter_id: str = ""
    name: str = ""
    father_husband_name: str = ""
    house_number: str = ""
    age: str = ""
    gender: str = ""
    part_no: str = ""
    ward_no: str = ""
    street_area_name: str = ""
    constituency_no: str = ""
    constituency_name: str = ""

    def is_valid(self) -> bool:
        return bool(self.name.strip())


# ── Summary page parser ────────────────────────────────────────────────────────

def parse_summary_count(pdf_path: str, total_pages: int, poppler_path) -> Optional[int]:
    log.info("Reading summary page to get expected voter count …")
    try:
        images = convert_from_path(
            pdf_path, dpi=DPI_DEFAULT,
            first_page=total_pages, last_page=total_pages,
            fmt="jpeg", poppler_path=poppler_path,
        )
        if not images:
            return None
        text = pytesseract.image_to_string(images[0], config=TESSERACT_CONFIG)
        patterns = [
            r"Total\s+Electors?\s*[:\-=]\s*([\d,]+)",
            r"Total\s+Voters?\s*[:\-=]\s*([\d,]+)",
            r"(?:Grand\s+)?Total\s*[:\-=]\s*([\d,]+)",
            r"மொத்தம்\s*[:\-=]\s*([\d,]+)",
            r"Total\s+\d+\s+\d+\s+\d+\s+([\d,]+)",
        ]
        for pat in patterns:
            m = re.search(pat, text, re.IGNORECASE)
            if m:
                count = int(m.group(1).replace(",", ""))
                log.info(f"Summary page → expected voter count: {count}")
                return count
        log.warning("Could not find voter count on summary page; validation skipped.")
        log.debug(f"Summary OCR text:\n{text[:800]}")
        return None
    except Exception as e:
        log.warning(f"Summary page read failed: {e}")
        return None


# ── Header parser ──────────────────────────────────────────────────────────────

def _extract_ward_no(text: str) -> str:
    """
    Extract ward number from any of the common Tamil Nadu roll formats:
      • "Ward No. : 5"
      • "Ward No : 5"
      • "Ward No 5"
      • "Ward : 5"
      • "Ward 5"
      • embedded in section name: "… Ward 12, …"  or  "Ward No.12"
    Returns the numeric (or alphanumeric) ward identifier, or "" if not found.
    """
    # Priority 1 — explicit "Ward No" with separator
    m = re.search(
        r"\bWard\s+No\.?\s*[:\-=\s]\s*(\w+)",
        text, re.IGNORECASE
    )
    if m:
        return m.group(1).strip()

    # Priority 2 — plain "Ward :" or "Ward =" with separator
    m = re.search(
        r"\bWard\s*[:\-=]\s*(\w+)",
        text, re.IGNORECASE
    )
    if m:
        return m.group(1).strip()

    # Priority 3 — "Ward" followed directly by a number (no separator)
    m = re.search(
        r"\bWard\s+(\d+)\b",
        text, re.IGNORECASE
    )
    if m:
        return m.group(1).strip()

    return ""


def parse_page_header(lines: list) -> dict:
    ctx = {
        "constituency_no": "", "constituency_name": "",
        "part_no": "", "section_no": "",
        "ward_no": "", "street_area_name": "",
    }
    full_header_text = " ".join(lines[:12])  # wider window for ward search

    for line in lines[:8]:
        m = re.search(
            r"Assembly\s+Constituency\s+No\s+and\s+Name\s*[:=]\s*"
            r"(\d+)[^\w]+([\w\s]+?)\s+Part\s+No\.?\s*[:\s]*(\d+)",
            line, re.IGNORECASE
        )
        if m:
            ctx["constituency_no"]   = m.group(1).strip()
            ctx["constituency_name"] = m.group(2).strip()
            ctx["part_no"]           = m.group(3).strip()

        m2 = re.search(r"Part\s+No\.?\s*[:\s]*(\d+)", line, re.IGNORECASE)
        if m2 and not ctx["part_no"]:
            ctx["part_no"] = m2.group(1).strip()

        m3 = re.search(r"Section\s+No\s+and\s+Name\s+(\d+)[^\w]+(.*)", line, re.IGNORECASE)
        if m3:
            ctx["section_no"]       = m3.group(1).strip()
            raw                     = m3.group(2).strip()
            ctx["street_area_name"] = raw
            # Try to pull ward from within the section name string
            ward_in_section = _extract_ward_no(raw)
            if ward_in_section:
                ctx["ward_no"] = ward_in_section

    # If ward still not found, scan all header lines directly
    if not ctx["ward_no"]:
        ward_from_header = _extract_ward_no(full_header_text)
        if ward_from_header:
            ctx["ward_no"] = ward_from_header

    return ctx


# ── Field line splitters ───────────────────────────────────────────────────────

def split_name_line(line: str) -> list:
    all_matches = list(re.finditer(
        r"\bName\s*[:\-!=+}\|;*?#\ufffd]|\bName\s+[a-z](?=\s)",
        line, re.IGNORECASE
    ))
    if not all_matches:
        return []
    filtered = []
    for m in all_matches:
        before = line[:m.start()].rstrip()
        if re.search(r"(?:Father|Husband|Mother)\s*$", before, re.IGNORECASE):
            continue
        filtered.append(m)
    if not filtered:
        return []
    results = []
    for i, m in enumerate(filtered):
        start = m.end()
        end   = filtered[i + 1].start() if i + 1 < len(filtered) else len(line)
        val   = line[start:end].strip().rstrip("-~").strip()
        val   = re.sub(r"\s+Photo\s*$", "", val, flags=re.IGNORECASE).strip()
        val   = re.sub(r"\s+(?:Father|Husband|Mother)$", "", val, flags=re.IGNORECASE).strip()
        if val:
            results.append(val)
    return results


def split_relation_line(line: str) -> list:
    matches = list(re.finditer(
        r"(?:Father|Husband|Mother)\s+(?:Name|Narne)\s*[:\-!=?+]\s*",
        line, re.IGNORECASE
    ))
    if not matches:
        return []
    results = []
    for i, m in enumerate(matches):
        start = m.end()
        end   = matches[i + 1].start() if i + 1 < len(matches) else len(line)
        val   = line[start:end].strip().rstrip("-~").strip()
        val   = re.sub(r"\s+Photo\s*$", "", val, flags=re.IGNORECASE).strip()
        if val:
            results.append(val)
    return results


def split_house_line(line: str) -> list:
    matches = list(re.finditer(r"House\s+Number\s*[:\-!=+?]\s*", line, re.IGNORECASE))
    if not matches:
        return []
    results = []
    for i, m in enumerate(matches):
        start = m.end()
        end   = matches[i + 1].start() if i + 1 < len(matches) else len(line)
        val   = line[start:end].strip()
        val   = re.split(r"\s+Photo\b|\s{3,}", val)[0].strip()
        if val:
            results.append(val)
    return results


def split_age_gender_line(line: str) -> list:
    def _norm_gender(g: str) -> str:
        g  = g.strip().title()
        lg = g.lower()
        if lg in ("femaie", "femail", "fernale"):
            return "Female"
        if lg in ("mate", "mal", "mae"):
            return "Male"
        if "third" in lg:
            return "Third Gender"
        return g

    results = []

    p1 = re.compile(
        r"(?:Age|A9e|Ag[e3])\s*[:\-!+?=]?\s*(\d{1,3}\.?)\s+"
        r"(?:Gender|Sex|Gen[d]er)\s*[:\-!+?=]\s*"
        r"(Male|Mate|Mae|Mal|Female|Femaie|Femail|Fernale|Third\s*Gender)",
        re.IGNORECASE
    )
    for m in p1.finditer(line):
        results.append((m.group(1).rstrip("."), _norm_gender(m.group(2))))

    if results:
        return results

    ages    = re.findall(r"(?:Age|A9e)\s*[:\-!+?=]?\s*(\d{1,3})", line, re.IGNORECASE)
    genders = re.findall(
        r"(?:Gender|Sex)\s*[:\-!+?=]\s*(Male|Mate|Mae|Mal|Female|Femaie|Femail|Third\s*Gender)",
        line, re.IGNORECASE
    )
    if ages and genders:
        for i in range(min(len(ages), len(genders))):
            results.append((ages[i], _norm_gender(genders[i])))
        return results

    age_only = re.findall(r"(?:Age|A9e)\s*[:\-!+?=]?\s*(\d{1,3})", line, re.IGNORECASE)
    if age_only:
        return [(a, "") for a in age_only]

    return []


def extract_voter_ids_from_serial_line(line: str) -> list:
    pattern = re.compile(r"(?i)([A-Za-z]{2,4})([0-9]{5,10})")
    return [(m.group(1) + m.group(2)).upper() for m in pattern.finditer(line)]


def extract_serials_from_line(line: str) -> list:
    return re.findall(r"(?:^|[|\[\(#\s])(\d{1,4})(?:[\]\)]|\s|$)", line)


def is_serial_line(line: str) -> bool:
    blocked = ["name", "father", "husband", "mother", "age", "gender",
               "house", "available", "photo", "assembly", "section",
               "constituency", "part no", "ward", "age as on"]
    ll = line.lower()
    if any(b in ll for b in blocked):
        return False
    if not re.search(r"\d", line):
        return False
    return bool(
        re.search(r"[A-Za-z]{2,}\d{3,}", line) or
        re.search(r"\[\d+\]", line) or
        re.search(r"^\s*\d{1,3}\s*$", line)
    )


# ── Group parser ───────────────────────────────────────────────────────────────

def _parse_group(grp_lines: list, ctx: dict) -> list:
    voter_ids, serials     = [], []
    names, fathers, houses = [], [], []
    age_genders: list      = []

    for gl in grp_lines:
        gll = gl.lower()
        if is_serial_line(gl):
            voter_ids = extract_voter_ids_from_serial_line(gl)
            serials   = extract_serials_from_line(gl)
            continue

        if "father name" in gll or "husband name" in gll or "mother name" in gll:
            f = split_relation_line(gl)
            if f:
                fathers = f

        if re.search(r"\bname\s*[:\-!=+}\|;*?#\ufffd]|\bname\s+[a-z]\s", gll):
            n = split_name_line(gl)
            if n:
                names = n

        if "house number" in gll:
            h = split_house_line(gl)
            if h:
                houses = h

        if re.search(r"\bage\b|\ba9e\b", gll):
            ag = split_age_gender_line(gl)
            if ag:
                age_genders.extend(ag)

    seen = set()
    deduped_ag = []
    for entry in age_genders:
        if entry not in seen:
            seen.add(entry)
            deduped_ag.append(entry)
    age_genders = deduped_ag

    records = []
    n_cards = max(len(names), len(voter_ids), 1)
    for k in range(min(n_cards, 3)):
        rec = VoterRecord(
            part_no=ctx.get("part_no", ""),
            ward_no=ctx.get("ward_no", ""),
            street_area_name=ctx.get("street_area_name", ""),
            constituency_no=ctx.get("constituency_no", ""),
            constituency_name=ctx.get("constituency_name", ""),
        )
        rec.sno                 = serials[k]   if k < len(serials)   else ""
        rec.voter_id            = voter_ids[k] if k < len(voter_ids) else ""
        rec.name                = names[k]     if k < len(names)     else ""
        rec.father_husband_name = fathers[k]   if k < len(fathers)   else ""
        rec.house_number        = houses[k]    if k < len(houses)    else ""
        if k < len(age_genders):
            rec.age, rec.gender = age_genders[k]
        if rec.is_valid():
            records.append(rec)
    return records


# ── Page processor ─────────────────────────────────────────────────────────────

def process_page(image, page_num: int, tess_config: str = TESSERACT_CONFIG) -> tuple:
    log.info(f"  OCR page {page_num} …")
    text  = pytesseract.image_to_string(image, config=tess_config)
    lines = [l.strip() for l in text.splitlines()]

    non_empty = [l for l in lines if l]
    ctx       = parse_page_header(non_empty[:12])
    log.info(
        f"    → {ctx['constituency_no']}-{ctx['constituency_name']} "
        f"| Part {ctx['part_no']} | Ward: {ctx['ward_no']}"
    )

    if not any(kw in text for kw in ["Name :", "Name:", "Age :", "Age:", "Gender :"]):
        log.info("    → Skip (no voter fields detected)")
        return ctx, []

    header_keywords = ["assembly constituency", "section no and name", "age as on"]
    current_ctx     = dict(ctx)
    records         = []
    current_group   = []

    for line in non_empty:
        ll = line.lower()

        if "section no and name" in ll:
            new_ctx = parse_page_header([line])
            for k, v in new_ctx.items():
                if v:
                    current_ctx[k] = v
            continue

        if any(kw in ll for kw in header_keywords):
            continue

        if re.search(r'\bAvailable\b', line, re.IGNORECASE):
            if re.search(r'\bAge\b.*\bGender\b', line, re.IGNORECASE):
                current_group.append(line)
            if current_group:
                records.extend(_parse_group(current_group, dict(current_ctx)))
                current_group = []
            continue

        current_group.append(line)

    if current_group:
        records.extend(_parse_group(current_group, dict(current_ctx)))

    missing = [r for r in records if not r.age or not r.gender]
    if missing:
        log.info(f"    → {len(missing)} records missing age/gender; running targeted re-OCR …")
        records = _recover_age_gender(image, records, tess_config)

    log.info(f"    → {len(records)} records extracted")
    return ctx, records


def _recover_age_gender(image, records: list, tess_config: str) -> list:
    alt_text = pytesseract.image_to_string(image, config=TESSERACT_CONFIG_ALT)
    all_ag   = []
    for line in alt_text.splitlines():
        ag = split_age_gender_line(line.strip())
        all_ag.extend(ag)

    missing_idx = [i for i, r in enumerate(records) if not r.age or not r.gender]
    fill_ag     = [(a, g) for a, g in all_ag if a]

    if not fill_ag:
        return records

    for j, idx in enumerate(missing_idx):
        if j < len(fill_ag):
            age, gender = fill_ag[j]
            if not records[idx].age:
                records[idx].age = age
            if not records[idx].gender:
                records[idx].gender = gender

    return records


# ── Validation + retry ─────────────────────────────────────────────────────────

def run_extraction(pdf_path: str, start_page: int, end_page: int, dpi: int,
                   poppler_path, tess_config: str = TESSERACT_CONFIG) -> tuple:
    log.info(f"Rendering pages {start_page}–{end_page} at DPI={dpi} …")
    images = convert_from_path(
        str(pdf_path), dpi=dpi,
        first_page=start_page, last_page=end_page,
        fmt="jpeg", thread_count=2,
        poppler_path=poppler_path,
    )
    log.info(f"Rendered {len(images)} page(s)")

    all_records  = []
    section_map: dict = {}
    current_ctx: dict = {}

    for idx, image in enumerate(images):
        ctx, records = process_page(image, start_page + idx, tess_config)
        for k, v in ctx.items():
            if v:
                current_ctx[k] = v

        for rec in records:
            if not rec.constituency_no:   rec.constituency_no   = current_ctx.get("constituency_no", "")
            if not rec.constituency_name: rec.constituency_name = current_ctx.get("constituency_name", "")
            if not rec.part_no:           rec.part_no           = current_ctx.get("part_no", "")
            if not rec.ward_no:           rec.ward_no           = current_ctx.get("ward_no", "")
            if not rec.street_area_name:  rec.street_area_name  = current_ctx.get("street_area_name", "")

        all_records.extend(records)
        for rec in records:
            key       = rec.street_area_name or f"Section_{current_ctx.get('section_no', '?')}"
            sheet_key = (key[:28] + "..") if len(key) > 30 else key
            section_map.setdefault(sheet_key, []).append(rec)

    return all_records, section_map, current_ctx


# ── Output writers ─────────────────────────────────────────────────────────────

HEADER_COLS = [
    "S.NO", "VOTER_ID", "NAME", "FATHER_HUSBAND_NAME",
    "HOUSE_NUMBER", "AGE", "GENDER",
    "PART_NO", "WARD_NO", "STREET_AREA_NAME",
    "CONSTITUENCY_NO", "CONSTITUENCY_NAME",
]
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
ALT_FILL    = PatternFill("solid", fgColor="D6E4F0")


def _record_row(rec: VoterRecord) -> list:
    d = asdict(rec)
    return [
        d["sno"], d["voter_id"], d["name"], d["father_husband_name"],
        d["house_number"], d["age"], d["gender"],
        d["part_no"], d["ward_no"], d["street_area_name"],
        d["constituency_no"], d["constituency_name"],
    ]


def _write_xlsx_sheet(ws, records: list):
    ws.append(HEADER_COLS)
    for cell in ws[1]:
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for i, rec in enumerate(records, start=2):
        ws.append(_record_row(rec))
        if i % 2 == 0:
            for cell in ws[i]:
                cell.fill = ALT_FILL

    widths = {"A":6,"B":14,"C":28,"D":28,"E":14,"F":6,"G":10,"H":8,"I":18,"J":50,"K":14,"L":22}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"


def save_to_xlsx(all_records, section_map, output_path: str,
                 batch_sheet_name: str = None):
    """Write one sheet (all records) to xlsx. section_map is ignored."""
    wb     = openpyxl.Workbook()
    ws_all = wb.active
    ws_all.title = batch_sheet_name if batch_sheet_name else "All Records"
    _write_xlsx_sheet(ws_all, all_records)
    log.info(f"  Sheet '{ws_all.title}' → {len(all_records)} rows")
    wb.save(output_path)
    log.info(f"Excel saved → {output_path}")


# ── CLI ────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Extract voter data from Tamil Nadu Electoral Roll PDFs"
    )
    parser.add_argument("pdf_path")
    parser.add_argument("--start-page", type=int, default=3,
        help="First page (1-based). Default=3 skips cover + map.")
    parser.add_argument("--end-page", type=int, default=None,
        help="Last page (1-based). Default=second-to-last (skips summary page).")
    parser.add_argument("--dpi", type=int, default=DPI_DEFAULT,
        help=f"Rendering DPI (default={DPI_DEFAULT}). Higher = better but slower.")
    parser.add_argument("--output", default=None,
        help="Output .xlsx path")
    parser.add_argument("--no-retry", action="store_true",
        help="Disable automatic cascading retry when count mismatches > 2%%")
    parser.add_argument("--batch", action="store_true",
        help="Batch mode: name the main xlsx sheet after the PDF filename stem")
    args = parser.parse_args()

    pdf_path = Path(args.pdf_path)
    if not pdf_path.exists():
        log.error(f"Not found: {pdf_path}")
        sys.exit(1)

    from pypdf import PdfReader
    total_pages = len(PdfReader(str(pdf_path)).pages)
    start_page  = args.start_page
    end_page    = args.end_page or (total_pages - 1)
    xlsx_path   = args.output or str(pdf_path.with_suffix(".xlsx"))

    # In batch mode the sheet name = PDF filename stem (truncated to 31 chars)
    batch_sheet_name = None
    if args.batch:
        stem = re.sub(r"[\\/*?:\[\]]", "", pdf_path.stem)[:31]
        batch_sheet_name = stem

    log.info(f"PDF: {pdf_path.name}  ({total_pages} pages)")
    log.info(f"Processing pages {start_page}–{end_page}")
    if batch_sheet_name:
        log.info(f"Batch mode — sheet name: '{batch_sheet_name}'")

    import os
    _poppler = r"C:\Users\sarat\AppData\Local\Microsoft\WinGet\Packages\oschwartz10612.Poppler_Microsoft.Winget.Source_8wekyb3d8bbwe\poppler-25.07.0\Library\bin"
    poppler_path = _poppler if os.path.isdir(_poppler) else None

    # ── Step 1: read summary to get expected count ──────────────────────────
    expected_count = parse_summary_count(str(pdf_path), total_pages, poppler_path)

    # ── Step 2: first extraction pass ──────────────────────────────────────
    all_records, section_map, current_ctx = run_extraction(
        str(pdf_path), start_page, end_page, args.dpi, poppler_path
    )
    log.info(f"\nFirst pass → {len(all_records)} records extracted")

    # ── Step 3: validate count; cascade retries until within 2% ────────────────
    if expected_count and not args.no_retry:
        best_records = all_records
        best_section = section_map
        best_gap     = abs(len(all_records) - expected_count)

        for attempt, (retry_dpi, retry_cfg) in enumerate(RETRY_SCHEDULE, start=2):
            extracted  = len(best_records)
            diff_ratio = abs(extracted - expected_count) / max(expected_count, 1)
            missed     = expected_count - extracted

            if diff_ratio <= MISMATCH_THRESHOLD:
                log.info(
                    f"Count OK after pass {attempt - 1}: "
                    f"{extracted}/{expected_count} ({diff_ratio:.2%} — within threshold)."
                )
                break

            log.warning(
                f"Pass {attempt - 1}: extracted {extracted} / expected {expected_count} "
                f"({diff_ratio:.2%} off = ~{missed} voters missing). "
                f"Starting pass {attempt} at DPI={retry_dpi}, config='{retry_cfg}' …"
            )

            retry_records, retry_section_map, _ = run_extraction(
                str(pdf_path), start_page, end_page,
                retry_dpi, poppler_path, tess_config=retry_cfg
            )
            retry_count = len(retry_records)
            log.info(f"Pass {attempt} → {retry_count} records")

            merged_by_id: dict = {}
            no_id_records: list = []

            def _absorb(records):
                for r in records:
                    if r.voter_id:
                        existing = merged_by_id.get(r.voter_id)
                        if existing is None:
                            merged_by_id[r.voter_id] = r
                        else:
                            if not existing.age    and r.age:    existing.age    = r.age
                            if not existing.gender and r.gender: existing.gender = r.gender
                            if not existing.name   and r.name:   existing.name   = r.name
                            if not existing.father_husband_name and r.father_husband_name:
                                existing.father_husband_name = r.father_husband_name
                    else:
                        no_id_records.append(r)

            _absorb(best_records)
            _absorb(retry_records)

            if len(retry_records) > len(best_records):
                no_id_from = [r for r in retry_records if not r.voter_id]
            else:
                no_id_from = [r for r in best_records  if not r.voter_id]

            merged     = list(merged_by_id.values()) + no_id_from
            merged_gap = abs(len(merged) - expected_count)

            log.info(
                f"After merge: {len(merged)} records "
                f"(gap {merged_gap} vs expected {expected_count})"
            )

            if merged_gap <= best_gap:
                best_gap     = merged_gap
                best_records = merged
                best_section = {}
                for rec in merged:
                    key       = rec.street_area_name or f"Section_{current_ctx.get('section_no', '?')}"
                    sheet_key = (key[:28] + "..") if len(key) > 30 else key
                    best_section.setdefault(sheet_key, []).append(rec)
            else:
                log.info("Merged result is worse than best so far; keeping previous best.")

        else:
            final_gap   = abs(len(best_records) - expected_count)
            final_ratio = final_gap / max(expected_count, 1)
            log.warning(
                f"All {len(RETRY_SCHEDULE) + 1} passes done. "
                f"Best result: {len(best_records)}/{expected_count} "
                f"({final_ratio:.2%} gap). Saving best available data."
            )

        all_records = best_records
        section_map = best_section

    elif expected_count and args.no_retry:
        diff_ratio = abs(len(all_records) - expected_count) / max(expected_count, 1)
        log.info(
            f"Retry disabled. Extracted {len(all_records)}/{expected_count} "
            f"({diff_ratio:.2%} gap)."
        )
    else:
        log.info("No expected count found on summary page; skipping validation.")

    final_count    = len(all_records)
    age_missing    = sum(1 for r in all_records if not r.age)
    gender_missing = sum(1 for r in all_records if not r.gender)
    log.info(
        f"\nFinal totals: {final_count} records | "
        f"{len(section_map)} sections | "
        f"missing age: {age_missing} | missing gender: {gender_missing}"
    )

    if not all_records:
        log.warning("No records extracted. Check --start-page / --end-page.")
        sys.exit(1)

    # ── Step 4: save xlsx only (no zip) ────────────────────────────────────
    save_to_xlsx(all_records, section_map, xlsx_path,
                 batch_sheet_name=batch_sheet_name)

    log.info(f"\n✅  Done")
    log.info(f"   Excel  → {xlsx_path}")


if __name__ == "__main__":
    main()