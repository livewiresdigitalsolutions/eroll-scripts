"""
Flask backend for the Electoral Roll Extractor UI.

Run with:
    pip install flask flask-cors
    python server.py

Flow per PDF:
  1. UI sends PDF + options + batch_name via POST /process
  2. Script runs → writes a single-sheet temp xlsx
  3. Server copies that sheet (named after the PDF stem) into {batch_name}.xlsx
  4. UI downloads the combined batch xlsx via GET /download/{batch_name}.xlsx
"""

import os
import re
import subprocess
import tempfile
from copy import copy
from pathlib import Path

import openpyxl
from flask import Flask, jsonify, request, send_file
from flask_cors import CORS

app = Flask(__name__)
CORS(app)

OUTPUT_DIR = Path(__file__).parent / "outputs"
OUTPUT_DIR.mkdir(exist_ok=True)

SCRIPT = Path(__file__).parent / "electoral_roll.py"


def _parse_record_count(log_text: str) -> int:
    m = re.search(r"Final totals:\s*(\d+)\s*records", log_text)
    if m:
        return int(m.group(1))
    matches = re.findall(r"→\s*(\d+)\s*records\s*extracted", log_text)
    return int(matches[-1]) if matches else 0


def _safe_sheet_name(name: str, existing: list) -> str:
    """Return a unique, xlsx-legal sheet name (max 31 chars)."""
    safe = re.sub(r"[\\/*?:\[\]]", "", name)[:31]
    base, idx = safe, 2
    while base in existing:
        base = safe[:28] + f"_{idx}"
        idx += 1
    return base


def _merge_sheet_into_batch(src_xlsx: Path, batch_xlsx: Path, sheet_name: str):
    """Copy the single sheet from src_xlsx into batch_xlsx as a new named sheet."""
    src_wb = openpyxl.load_workbook(src_xlsx)
    src_ws = src_wb.active

    if batch_xlsx.exists():
        dst_wb = openpyxl.load_workbook(batch_xlsx)
    else:
        dst_wb = openpyxl.Workbook()
        dst_wb.remove(dst_wb.active)  # drop the blank default sheet

    name = _safe_sheet_name(sheet_name, dst_wb.sheetnames)
    dst_ws = dst_wb.create_sheet(title=name)

    for row in src_ws.iter_rows():
        for cell in row:
            dst_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                dst_cell.font      = copy(cell.font)
                dst_cell.fill      = copy(cell.fill)
                dst_cell.alignment = copy(cell.alignment)

    for col, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col].width = dim.width

    if src_ws.freeze_panes:
        dst_ws.freeze_panes = src_ws.freeze_panes

    dst_wb.save(batch_xlsx)


@app.route("/health")
def health():
    return jsonify({"ok": True})


@app.route("/process", methods=["POST"])
def process():
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "No file provided"}), 400

    pdf_file = request.files["file"]
    if not pdf_file.filename.lower().endswith(".pdf"):
        return jsonify({"ok": False, "error": "Only PDF files accepted"}), 400

    start_page = request.form.get("start_page", "3")
    end_page   = request.form.get("end_page", "").strip()
    dpi        = request.form.get("dpi", "200")
    no_retry   = request.form.get("no_retry", "false").lower() == "true"
    batch_name = re.sub(r"[^\w\-]", "_", request.form.get("batch_name", "batch").strip()) or "batch"

    pdf_stem   = Path(pdf_file.filename).stem
    batch_xlsx = OUTPUT_DIR / f"{batch_name}.xlsx"

    # Save uploaded PDF to a temp file
    tmp_fd, tmp_pdf = tempfile.mkstemp(suffix=".pdf", prefix=f"{pdf_stem}_")
    os.close(tmp_fd)

    # Script writes a single-sheet temp xlsx, which we then merge into the batch file
    tmp_fd2, tmp_xlsx = tempfile.mkstemp(suffix=".xlsx", prefix=f"{pdf_stem}_")
    os.close(tmp_fd2)

    try:
        pdf_file.save(tmp_pdf)

        cmd = [
            "python", str(SCRIPT),
            tmp_pdf,
            "--start-page", start_page,
            "--dpi", dpi,
            "--output", tmp_xlsx,
            "--batch",
        ]
        if end_page:
            cmd += ["--end-page", end_page]
        if no_retry:
            cmd += ["--no-retry"]

        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            cwd=str(Path(__file__).parent),
        )

        log_text = result.stdout + result.stderr
        records  = _parse_record_count(log_text)
        script_ok = result.returncode == 0 and Path(tmp_xlsx).exists()

        if script_ok:
            _merge_sheet_into_batch(Path(tmp_xlsx), batch_xlsx, pdf_stem)

        return jsonify({
            "ok":         script_ok,
            "records":    records,
            "batch_xlsx": batch_xlsx.name if script_ok else None,
            "log":        log_text,
            "error":      None if script_ok else f"Script exited with code {result.returncode}",
        })

    finally:
        for f in (tmp_pdf, tmp_xlsx):
            if os.path.exists(f):
                os.unlink(f)


@app.route("/download/<path:filename>")
def download(filename):
    safe = Path(filename).name
    path = OUTPUT_DIR / safe
    if not path.exists():
        return jsonify({"error": "File not found"}), 404
    return send_file(str(path), as_attachment=True)


if __name__ == "__main__":
    print("Electoral Roll backend listening on http://localhost:5000")
    print(f"Outputs folder: {OUTPUT_DIR}")
    app.run(port=5000, debug=False)
